"""
exportObservaciones.py
Genera Excel de observaciones con formato corporativo usando openpyxl.
Estructura: hoja "Todas" + una hoja por tipo de activo.
Columnas humanas por defecto, columnas técnicas opcionales.

Ubicación: backend/src/utils/exportObservaciones.py
Uso: python3 exportObservaciones.py <payload.json> <output.xlsx>
"""

import sys
import json
import copy
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# ── Colores corporativos ──
COLOR_HEADER_BG  = "861F41"   # rojo oscuro — fondo header columnas
COLOR_HEADER_FG  = "FFFFFF"   # blanco — texto header
COLOR_TITLE_BG   = "FA8200"   # naranja — fila título del reporte
COLOR_TITLE_FG   = "FFFFFF"
COLOR_ALT_ROW    = "FDF0E8"   # naranja muy claro — filas alternas
COLOR_BORDER     = "D0D0D0"   # gris claro — bordes

# ── Colores por tipo de evento ──
EVENTO_COLORS = {
    "Nota":            "FFF3E0",   # naranja muy claro
    "Mantenimiento":   "E8F5E9",   # verde claro
    "Incidente":       "FFEBEE",   # rojo muy claro
    "Cambio de campo": "F3F0FF",   # morado muy claro
    "Importación":     "F5F5F5",   # gris claro
}

EVENTO_ACCENT = {
    "Nota":            "FA8200",
    "Mantenimiento":   "2E7D32",
    "Incidente":       "C62828",
    "Cambio de campo": "6A1B9A",
    "Importación":     "888888",
}

def thin_side(color="D0D0D0"):
    return Side(style='thin', color=color)

def cell_border(color="D0D0D0"):
    s = thin_side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font():
    return Font(name='Calibri', bold=True, size=11, color="FFFFFF")

def data_font(color="000000", bold=False):
    return Font(name='Calibri', size=10, color=color, bold=bold)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=False)

def left_wrap():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=False)


# ── Insertar logo ──
def add_logo_to_sheet(ws, logo_filename='logo.png'):
    """Agrega logo en B2 — tamaño 236x51px (modo seguro, sin romper Excel)."""
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        logo_path = os.path.join(script_dir, logo_filename)
        
        # Verificar que el archivo existe Y es un archivo válido
        if not os.path.exists(logo_path) or os.path.getsize(logo_path) < 100:
            return  # Archivo no existe o es muy pequeño
        
        img = XLImage(logo_path)
        img.width = 236
        img.height = 51
        ws.add_image(img, 'B2')
    except Exception as e:
        # Silenciar CUALQUIER error — la imagen es opcional
        pass


# ── Definición de columnas ──
# incluir_tecnicos: si True agrega Campo modificado, Valor anterior, Valor nuevo
def get_columns(incluir_tecnicos=False, omit_codigo=False):
    cols = [
        {"key": "Activo",         "header": "Activo",         "width": 30, "wrap": False},
        {"key": "Tipo",           "header": "Tipo",           "width": 16, "wrap": False},
    ]
    if not omit_codigo:
        cols.append({"key": "Código", "header": "Código", "width": 18, "wrap": False})
    cols += [
        {"key": "Fecha",          "header": "Fecha",          "width": 22, "wrap": False},
        {"key": "Autor",          "header": "Autor",          "width": 20, "wrap": False},
        {"key": "Tipo de evento", "header": "Tipo de Evento", "width": 18, "wrap": False},
        {"key": "Descripción",    "header": "Descripción",    "width": 60, "wrap": True},
    ]
    if incluir_tecnicos:
        cols += [
            {"key": "Campo modificado", "header": "Campo Modificado", "width": 22, "wrap": False},
            {"key": "Valor anterior",   "header": "Valor Anterior",   "width": 24, "wrap": False},
            {"key": "Valor nuevo",      "header": "Valor Nuevo",      "width": 24, "wrap": False},
        ]
    return cols


def escribir_hoja(wb, nombre_hoja, rows, incluir_tecnicos=False, omit_codigo=False):
    """Crea y formatea una hoja con las observaciones dadas."""
    ws = wb.create_sheet(title=nombre_hoja[:31])
    add_logo_to_sheet(ws)  # ✅ Agregar logo en esquina izquierda

    cols = get_columns(incluir_tecnicos, omit_codigo=omit_codigo)
    num_cols = len(cols)

    # ── Fila 1: título del reporte ──
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Observaciones de Bitácora — {nombre_hoja}"
    title_cell.font = Font(name='Calibri', bold=True, size=13, color=COLOR_TITLE_FG)
    title_cell.fill = fill(COLOR_TITLE_BG)
    title_cell.alignment = center()
    title_cell.border = cell_border("FA8200")

    # Aplicar fill a todas las celdas de la fila de título
    for c in range(2, num_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill(COLOR_TITLE_BG)
        cell.border = cell_border("FA8200")

    # ── Fila 2: subtítulo con conteo ──
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = f"Total de registros: {len(rows)}  |  Exportado: {__import__('datetime').datetime.now().strftime('%d/%m/%Y %H:%M')}"
    sub_cell.font = Font(name='Calibri', size=9, color="555555", italic=True)
    sub_cell.fill = fill("FFF8F0")
    sub_cell.alignment = left()
    for c in range(2, num_cols + 1):
        ws.cell(row=2, column=c).fill = fill("FFF8F0")

    # ── Fila 3: headers de columnas ──
    ws.row_dimensions[3].height = 22
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=3, column=ci)
        cell.value = col["header"]
        cell.font = header_font()
        cell.fill = fill(COLOR_HEADER_BG)
        cell.alignment = center()
        cell.border = cell_border("5A0000")
        ws.column_dimensions[get_column_letter(ci)].width = col["width"]

    # ── Filas de datos (desde fila 4) ──
    for ri, row in enumerate(rows):
        excel_row = ri + 4
        ws.row_dimensions[excel_row].height = 18

        evento = row.get("Tipo de evento", "")
        bg_color = EVENTO_COLORS.get(evento, "FFFFFF")
        accent   = EVENTO_ACCENT.get(evento, "000000")

        # Alternar: si el color del evento es el default, usar alt row
        if bg_color == "FFFFFF" and ri % 2 == 1:
            bg_color = COLOR_ALT_ROW

        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=excel_row, column=ci)
            val = row.get(col["key"], "")
            cell.value = val if val else None

            # Fuente especial para tipo de evento (color del acento)
            if col["key"] == "Tipo de evento":
                cell.font = Font(name='Calibri', size=10, color=accent, bold=True)
            elif col["key"] == "Activo":
                cell.font = Font(name='Calibri', size=10, color="1A1A1A", bold=True)
            else:
                cell.font = data_font()

            cell.fill = fill(bg_color)
            cell.border = cell_border()
            cell.alignment = left_wrap() if col.get("wrap") else left()

    # ── Freeze panes (header fijo) ──
    ws.freeze_panes = ws.cell(row=4, column=1)

    # ── Autofilter en headers ──
    ws.auto_filter.ref = f"A3:{get_column_letter(num_cols)}3"

    return ws


def main():
    payload_path = sys.argv[1]
    output_path  = sys.argv[2]

    with open(payload_path, 'r', encoding='utf-8') as f:
        payload = json.load(f)

    rows             = payload['rows']             # todas las filas
    incluir_tecnicos = payload.get('incluirTecnicos', False)

    if not rows:
        # Crear workbook vacío
        wb = Workbook()
        wb.active.title = "Sin datos"
        wb.save(output_path)
        return

    wb = Workbook()
    # Quitar hoja default
    wb.remove(wb.active)

    # ── Hoja "Todas" ──
    escribir_hoja(wb, "Todas", rows, incluir_tecnicos, omit_codigo=False)

    # ── Hojas por tipo ──
    singular_to_plural = {
        "Servidor":     "Servidores",
        "Base de Datos": "Bases de Datos",
        "Red":          "Red",
        "UPS":          "UPS",
        "VPN":          "VPN",
        "Móvil":        "Móviles",
    }

    # Agrupar por tipo
    grupos: dict = {}
    for row in rows:
        tipo = (row.get("Tipo") or "Otros").strip()
        if tipo not in grupos:
            grupos[tipo] = []
        grupos[tipo].append(row)

    for tipo_singular, lista in grupos.items():
        if not lista:
            continue
        nombre_hoja = singular_to_plural.get(tipo_singular, tipo_singular)
        omit_codigo = tipo_singular in {"Móvil", "UPS", "Base de Datos"}
        escribir_hoja(wb, nombre_hoja, lista, incluir_tecnicos, omit_codigo=omit_codigo)

    wb.save(output_path)
    print(f"OK: {output_path} — {len(rows)} filas")


if __name__ == '__main__':
    main()
    
    