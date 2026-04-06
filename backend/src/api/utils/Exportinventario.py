"""
exportInventario.py - VERSIÓN DEFINITIVA v4
+ Hoja InventarioVPN  (Nombre, Conexión, Fases, Origen, Destino)
+ Hoja InventarioMovil (todos los campos)
Sin cambios en las hojas existentes.

Ubicación: backend/src/api/utils/exportInventario.py
Uso:       python exportInventario.py <payload.json> <output.xlsx>
"""

import sys
import json
import copy
import re
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────
# PARSEO DE FECHAS
# ─────────────────────────────────────────────────────────────────────
_JS_MONTHS = {
    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4,  'May': 5,  'Jun': 6,
    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12,
}

def format_date(val):
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.strftime('%d/%m/%Y')
    s = str(val).strip()
    if not s:
        return None
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    m = re.search(r'(\w{3})\s+(\w{3})\s+(\d{1,2})\s+(\d{4})', s)
    if m:
        month_str = m.group(2)
        day   = int(m.group(3))
        year  = int(m.group(4))
        month = _JS_MONTHS.get(month_str)
        if month:
            return f"{day:02d}/{month:02d}/{year}"
    if re.match(r'^\d{2}/\d{2}/\d{4}$', s):
        return s
    return s

DATE_KEYS = {'fechaFinSoporte', 'fechaFinalSoporte', 'Fecha Entrega', 'Fecha Devolución'}

# ─────────────────────────────────────────────────────────────────────
# ESTILOS — iguales al template para hojas existentes
# ─────────────────────────────────────────────────────────────────────
def make_border(left='thin', right='thin', top='thin', bottom='thin'):
    def side(style):
        return Side(style=style) if style else Side(style=None)
    return Border(left=side(left), right=side(right),
                  top=side(top),   bottom=side(bottom))

BORDER_THIN  = make_border()
BORDER_UPS_B = make_border(left='medium')
BORDER_UPS_D = make_border(left='thin', right='medium')
BORDER_UPS   = make_border(left='medium')
BORDER_BD_B  = make_border(left='medium')
BORDER_BD_N  = make_border(right='medium')
BORDER_BD    = make_border()

FONT_DATA  = Font(name='Calibri', size=11, bold=False, color='FF000000')
ALIGN_DATA = Alignment(wrap_text=True, vertical='center')

ROW_HEIGHTS = {
    'InventarioServidores': 15.75,
    'InventarioRedes':      30.75,
    'InventarioUPS':        30.75,
    'InventarioBD':         15.75,
    'InventarioVPN':        20.0,
    'InventarioMovil':      20.0,
}

def get_border(sheet_name, col_index_0based, total_cols):
    if sheet_name == 'InventarioUPS':
        if col_index_0based == 0:              return copy.copy(BORDER_UPS_B)
        if col_index_0based == 2:              return copy.copy(BORDER_UPS_D)
        return copy.copy(BORDER_UPS)
    if sheet_name == 'InventarioBD':
        if col_index_0based == 0:              return copy.copy(BORDER_BD_B)
        if col_index_0based == total_cols - 1: return copy.copy(BORDER_BD_N)
        return copy.copy(BORDER_BD)
    return copy.copy(BORDER_THIN)

# ─────────────────────────────────────────────────────────────────────
# COPIA DE HEADER (filas 1-10 del template)
# ─────────────────────────────────────────────────────────────────────
def copy_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        dst.font          = copy.copy(src.font)
        dst.border        = copy.copy(src.border)
        dst.fill          = copy.copy(src.fill)
        dst.number_format = src.number_format
        dst.alignment     = copy.copy(src.alignment)

def copy_header(src_ws, dst_ws, header_rows=10):
    for row in src_ws.iter_rows(min_row=1, max_row=header_rows):
        for cell in row:
            copy_cell(cell, dst_ws.cell(row=cell.row, column=cell.column))
    for merge in src_ws.merged_cells.ranges:
        if merge.max_row <= header_rows:
            dst_ws.merge_cells(str(merge))
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width
    for row_idx in range(1, header_rows + 1):
        rd = src_ws.row_dimensions.get(row_idx)
        if rd and rd.height:
            dst_ws.row_dimensions[row_idx].height = rd.height

# ─────────────────────────────────────────────────────────────────────
# DATOS EN HOJAS CON TEMPLATE (sin cambios)
# ─────────────────────────────────────────────────────────────────────
def write_data_rows(ws, sheet_name, data_rows, col_map, date_cols=None, start_row=11):
    if date_cols is None:
        date_cols = set()
    row_height = ROW_HEIGHTS.get(sheet_name, 15.75)
    total_cols = len(col_map)
    for i, row_data in enumerate(data_rows):
        row_num = start_row + i
        for j, key in enumerate(col_map):
            col_num = 2 + j
            raw_val = row_data.get(key, None)
            if j in date_cols or key in DATE_KEYS:
                val = format_date(raw_val)
            else:
                val = raw_val if raw_val not in ('', None) else None
            cell           = ws.cell(row=row_num, column=col_num)
            cell.value     = val
            cell.font      = copy.copy(FONT_DATA)
            cell.border    = get_border(sheet_name, j, total_cols)
            cell.alignment = copy.copy(ALIGN_DATA)
        ws.row_dimensions[row_num].height = row_height

# ─────────────────────────────────────────────────────────────────────
# HOJAS EXTRA — VPN y Móvil (header corporativo propio)
# ─────────────────────────────────────────────────────────────────────
def crear_hoja_extra(dst_wb, sheet_name, labels, widths, data_rows, date_label_keys=None):
    """
    labels         : lista de strings — nombre visible de cada columna
    widths         : lista de anchos (mismo largo que labels)
    data_rows      : lista de dicts donde las KEYS son los mismos strings que labels
    date_label_keys: set de labels que son fechas → se formatean DD/MM/YYYY
    """
    if date_label_keys is None:
        date_label_keys = set()

    ws       = dst_wb.create_sheet(sheet_name)
    num_cols = len(labels)

    # Fila 1 — título naranja
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    titulo       = ws.cell(row=1, column=1)
    titulo.value = sheet_name
    titulo.font  = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    titulo.fill  = PatternFill('solid', fgColor='FA8200')
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    for c in range(2, num_cols + 1):
        ws.cell(row=1, column=c).fill = PatternFill('solid', fgColor='FA8200')
    ws.row_dimensions[1].height = 24

    # Fila 2 — headers rojo oscuro
    h_font   = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    h_fill   = PatternFill('solid', fgColor='861F41')
    h_align  = Alignment(horizontal='center', vertical='center', wrap_text=False)
    h_border = make_border()
    for ci, (label, width) in enumerate(zip(labels, widths), start=1):
        cell           = ws.cell(row=2, column=ci)
        cell.value     = label
        cell.font      = copy.copy(h_font)
        cell.fill      = copy.copy(h_fill)
        cell.alignment = copy.copy(h_align)
        cell.border    = copy.copy(h_border)
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[2].height = 18

    # Freeze en fila 3
    ws.freeze_panes = ws.cell(row=3, column=1)

    # Datos desde fila 3
    row_h      = ROW_HEIGHTS.get(sheet_name, 20.0)
    d_font     = Font(name='Calibri', size=11, color='FF000000')
    d_align    = Alignment(wrap_text=True, vertical='center', horizontal='left')
    d_border   = make_border()

    for ri, row_data in enumerate(data_rows):
        row_num = ri + 3
        for ci, label in enumerate(labels, start=1):
            raw_val = row_data.get(label, None)
            if label in date_label_keys:
                val = format_date(raw_val)
            else:
                val = raw_val if raw_val not in ('', None) else None
            cell           = ws.cell(row=row_num, column=ci)
            cell.value     = val
            cell.font      = copy.copy(d_font)
            cell.border    = copy.copy(d_border)
            cell.alignment = copy.copy(d_align)
        ws.row_dimensions[row_num].height = row_h

    return ws

# ─────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 3:
        print("Uso: python exportInventario.py <payload.json> <output.xlsx>")
        sys.exit(1)

    with open(sys.argv[1], 'r', encoding='utf-8') as f:
        payload = json.load(f)

    src_wb = load_workbook(payload['template'], read_only=False, data_only=True)
    dst_wb = Workbook()
    dst_wb.remove(dst_wb.active)

    # ── SERVIDORES ──────────────────────────────────────────────────
    src_ws = src_wb['InventarioServidores']
    dst_ws = dst_wb.create_sheet('InventarioServidores')
    copy_header(src_ws, dst_ws)
    write_data_rows(dst_ws, 'InventarioServidores', payload['servidores'], [
        'nombre', 'propietario', 'custodio', 'monitoreo', 'backup',
        'ipInterna', 'ipGestion', 'ipServicio',
        'ambiente', 'tipoServidor', 'appSoporta', 'ubicacion',
        'vcpu', 'vramMb', 'sistemaOperativo',
        'fechaFinSoporte', 'rutasBackup', 'contratoQueSoporta',
    ], date_cols={15})

    # ── REDES ────────────────────────────────────────────────────────
    src_ws = src_wb['InventarioRedes']
    dst_ws = dst_wb.create_sheet('InventarioRedes')
    copy_header(src_ws, dst_ws)
    write_data_rows(dst_ws, 'InventarioRedes', payload['redes'], [
        'nombre', 'propietario', 'custodio', 'serial', 'mac', 'modelo',
        'fechaFinSoporte', 'ipGestion', 'estado', 'codigoServicio',
        'ubicacion', 'contratoQueSoporta',
    ], date_cols={6})

    # ── UPS ──────────────────────────────────────────────────────────
    src_ws = src_wb['InventarioUPS']
    dst_ws = dst_wb.create_sheet('InventarioUPS')
    copy_header(src_ws, dst_ws)
    write_data_rows(dst_ws, 'InventarioUPS', payload['ups'], [
        'nombre', 'propietario', 'custodio', 'serial', 'placa',
        'modelo', 'estado', 'ubicacion',
    ])

    # ── BASES DE DATOS ───────────────────────────────────────────────
    src_ws = src_wb['InventarioBD']
    dst_ws = dst_wb.create_sheet('InventarioBD')
    copy_header(src_ws, dst_ws)
    write_data_rows(dst_ws, 'InventarioBD', payload['bds'], [
        'nombre', 'propietario', 'custodio', 'servidor1', 'servidor2',
        'racScan', 'ambiente', 'appSoporta', 'versionBd',
        'fechaFinalSoporte', 'contenedorFisico', 'contratoQueSoporta',
    ], date_cols={9})

    # ── VPN ──────────────────────────────────────────────────────────
    if payload.get('vpns'):
        vpn_labels = ['Nombre', 'Conexión', 'Fases', 'Origen', 'Destino']
        vpn_widths = [40, 24, 14, 50, 50]
        vpn_rows = [{
            'Nombre':   r.get('nombre', ''),
            'Conexión': r.get('conexion', ''),
            'Fases':    r.get('fases', ''),
            'Origen':   r.get('origen', ''),
            'Destino':  r.get('destino', ''),
        } for r in payload['vpns']]
        crear_hoja_extra(dst_wb, 'InventarioVPN', vpn_labels, vpn_widths, vpn_rows)

    # ── MÓVIL ────────────────────────────────────────────────────────
    if payload.get('moviles'):
        mov_labels = [
            'Nombre', '# Caso', 'Región', 'Dependencia', 'Sede', 'C.C.',
            'Usuario Red', 'Correo', 'UNI', 'Marca', 'Modelo', 'Serial',
            'IMEI 1', 'IMEI 2', 'SIM', 'Número Línea',
            'Fecha Entrega', 'Obs. Entrega', 'Fecha Devolución', 'Obs. Devolución',
        ]
        mov_widths = [
            26, 14, 22, 28, 18, 14,
            20, 34, 14, 16, 20, 20,
            18, 18, 20, 16,
            16, 40, 16, 40,
        ]
        mov_rows = [{
            'Nombre':           r.get('nombre', ''),
            '# Caso':           r.get('numeroCaso', ''),
            'Región':           r.get('region', ''),
            'Dependencia':      r.get('dependencia', ''),
            'Sede':             r.get('sede', ''),
            'C.C.':             r.get('cedula', ''),
            'Usuario Red':      r.get('usuarioRed', ''),
            'Correo':           r.get('correoResponsable', ''),
            'UNI':              r.get('uni', ''),
            'Marca':            r.get('marca', ''),
            'Modelo':           r.get('modelo', ''),
            'Serial':           r.get('serial', ''),
            'IMEI 1':           r.get('imei1', ''),
            'IMEI 2':           r.get('imei2', ''),
            'SIM':              r.get('sim', ''),
            'Número Línea':     r.get('numeroLinea', ''),
            'Fecha Entrega':    r.get('fechaEntrega', ''),
            'Obs. Entrega':     r.get('observacionesEntrega', ''),
            'Fecha Devolución': r.get('fechaDevolucion', ''),
            'Obs. Devolución':  r.get('observacionesDevolucion', ''),
        } for r in payload['moviles']]
        crear_hoja_extra(dst_wb, 'InventarioMovil', mov_labels, mov_widths, mov_rows,
                         date_label_keys={'Fecha Entrega', 'Fecha Devolución'})

    # ── CONTROL DE CAMBIOS ───────────────────────────────────────────
    if 'Control de Cambios' in src_wb.sheetnames:
        src_ws = src_wb['Control de Cambios']
        dst_ws = dst_wb.create_sheet('Control de Cambios')
        copy_header(src_ws, dst_ws)

    src_wb.close()
    dst_wb.save(sys.argv[2])
    print(f"OK: {sys.argv[2]}")

if __name__ == '__main__':
    main()